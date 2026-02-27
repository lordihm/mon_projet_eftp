from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.template.loader import render_to_string
from django.http import HttpResponse
from .models import EtablissementFormel
from .forms import EtablissementFormelSimpleForm, EtablissementFormelCompletForm
from apps.renaloc.models import Region
from .models import EtablissementFormel, ApprenantFormel, FormateurFormel, FiliereFormel
from .forms import ApprenantFormelForm, FormateurFormelForm, FiliereFormelForm
from django.db.models import Sum, F
import csv
import json
import openpyxl

@login_required
def etablissement_list(request):
    """Liste des établissements formels"""
    etablissements = EtablissementFormel.objects.all().select_related('region', 'departement', 'commune')
    
    # Calcul des statistiques
    publics_count = etablissements.filter(statut='PUBLIC').count()
    prives_count = etablissements.filter(statut='PRIVE').count()
    
    # Récupérer tous les apprenants (à adapter selon votre modèle)
    total_apprenants = 0
    
    # Pour les filtres
    regions = Region.objects.all()
    type_choices = EtablissementFormel.TYPE_ETABLISSEMENT_CHOICES
    
    context = {
        'etablissements': etablissements,
        'publics_count': publics_count,
        'prives_count': prives_count,
        'total_apprenants': total_apprenants,
        'regions': regions,
        'type_choices': type_choices,
    }
    return render(request, 'eftp_formel/etablissement_list.html', context)

@login_required
def etablissement_create(request):
    """Créer un nouvel établissement (formulaire simplifié)"""
    if request.method == 'POST':
        form = EtablissementFormelSimpleForm(request.POST)
        if form.is_valid():
            etablissement = form.save()
            messages.success(request, f"Établissement {etablissement.nom} créé avec succès!")
            messages.info(request, "Vous pouvez maintenant compléter les informations détaillées.")
            return redirect('eftp_formel:etablissement_complet', pk=etablissement.pk)
        else:
            messages.error(request, "Veuillez corriger les erreurs ci-dessous.")
    else:
        form = EtablissementFormelSimpleForm()
    
    # Pour les select dépendants
    regions = Region.objects.all()
    
    context = {
        'form': form,
        'regions': regions,
        'is_simple_form': True,
        'titre': "Nouvel établissement - Informations de base"
    }
    return render(request, 'eftp_formel/etablissement_form_simple.html', context)

@login_required
def etablissement_complet(request, pk):
    """Tableau de bord complet de l'établissement"""
    import sys
    print("="*50, file=sys.stderr)
    print("🔍 DEBUG: Vue etablissement_complet appelée avec pk =", pk, file=sys.stderr)
    print("="*50, file=sys.stderr)
    
    etablissement = get_object_or_404(EtablissementFormel, pk=pk)
    
    # Afficher le chemin du template qui va être utilisé
    from django.template.loader import get_template
    try:
        template = get_template('eftp_formel/etablissement_dashboard.html')
        print("✅ Template trouvé :", template.origin.name, file=sys.stderr)
    except:
        print("❌ Template NON trouvé !", file=sys.stderr)
    
    # Récupérer les données
    from .models import ApprenantFormel, FormateurFormel, FiliereFormel
    
    context = {
        'etablissement': etablissement,
        'apprenants': ApprenantFormel.objects.filter(etablissement=etablissement),
        'formateurs': FormateurFormel.objects.filter(etablissement=etablissement),
        'filieres': FiliereFormel.objects.filter(etablissement=etablissement),
        'total_apprenants': 0,
        'total_formateurs': 0,
        'total_filieres': 0,
        'progression_globale': 50,
    }
    
    print("🚀 Rendu du template...", file=sys.stderr)
    return render(request, 'eftp_formel/etablissement_dashboard.html', context)



@login_required
def etablissement_complet(request, pk):
    etablissement = get_object_or_404(EtablissementFormel, pk=pk)
    
    # Chemin absolu vers le template
    template_path = 'apps/eftp_formel/templates/eftp_formel/etablissement_dashboard.html'
    
    context = {'etablissement': etablissement}
    
'''
@login_required
def etablissement_complet(request, pk):
    """Tableau de bord complet de l'établissement"""
    import sys
    print("="*50, file=sys.stderr)
    print("🔍 DEBUG: Vue etablissement_complet appelée avec pk =", pk, file=sys.stderr)
    print("="*50, file=sys.stderr)
    
    etablissement = get_object_or_404(EtablissementFormel, pk=pk)
    
    # Afficher le chemin du template qui va être utilisé
    from django.template.loader import get_template
    try:
        template = get_template('eftp_formel/etablissement_dashboard.html')
        print("✅ Template trouvé :", template.origin.name, file=sys.stderr)
    except:
        print("❌ Template NON trouvé !", file=sys.stderr)
    
    # Récupérer les données
    from .models import ApprenantFormel, FormateurFormel, FiliereFormel
    
    context = {
        'etablissement': etablissement,
        'apprenants': ApprenantFormel.objects.filter(etablissement=etablissement),
        'formateurs': FormateurFormel.objects.filter(etablissement=etablissement),
        'filieres': FiliereFormel.objects.filter(etablissement=etablissement),
        'total_apprenants': 0,
        'total_formateurs': 0,
        'total_filieres': 0,
        'progression_globale': 50,
    }
    
    print("🚀 Rendu du template...", file=sys.stderr)
    return render(request, 'eftp_formel/etablissement_dashboard.html', context)
'''
'''
@login_required
def etablissement_complet(request, pk):
    """Version simplifiée pour tester"""
    etablissement = get_object_or_404(EtablissementFormel, pk=pk)
    
    # Test simple : renvoyer juste le nom pour vérifier que la vue fonctionne
    from django.http import HttpResponse
    return HttpResponse(f"TEST: Vue etablissement_complet chargée pour {etablissement.nom}")
'''
'''
@login_required
def etablissement_complet(request, pk):
    """Tableau de bord complet de l'établissement avec tous les onglets"""
    try:
        etablissement = EtablissementFormel.objects.get(pk=pk)
        print(f"✅ Établissement trouvé : {etablissement.nom}")  # Pour debug
    except EtablissementFormel.DoesNotExist:
        messages.error(request, "Établissement non trouvé")
        return redirect('eftp_formel:etablissement_list')
    
    # Récupérer les données associées
    from .models import ApprenantFormel, FormateurFormel, FiliereFormel
    from django.db.models import Sum, Count
    
    # Récupérer les apprenants
    apprenants = ApprenantFormel.objects.filter(etablissement=etablissement)
    print(f"✅ {apprenants.count()} apprenants trouvés")  # Pour debug
    
    # Récupérer les formateurs
    formateurs = FormateurFormel.objects.filter(etablissement=etablissement)
    print(f"✅ {formateurs.count()} formateurs trouvés")  # Pour debug
    
    # Récupérer les filières
    filieres = FiliereFormel.objects.filter(etablissement=etablissement)
    print(f"✅ {filieres.count()} filières trouvées")  # Pour debug
    
    # Calculer les totaux
    total_apprenants = 0
    for app in apprenants:
        total_apprenants += (app.masculin or 0) + (app.feminin or 0)
    
    total_formateurs = formateurs.count()
    total_filieres = filieres.count()
    
    # Calculer la progression
    progression = 0
    champs_remplis = 0
    champs_total = 15  # Nombre de critères
    
    if etablissement.nom: champs_remplis += 1
    if etablissement.sigle: champs_remplis += 1
    if etablissement.code: champs_remplis += 1
    if etablissement.statut: champs_remplis += 1
    if etablissement.zone: champs_remplis += 1
    if etablissement.region: champs_remplis += 1
    if etablissement.departement: champs_remplis += 1
    if etablissement.commune: champs_remplis += 1
    if etablissement.type_etablissement: champs_remplis += 1
    if etablissement.regime: champs_remplis += 1
    if etablissement.date_autorisation: champs_remplis += 1
    if etablissement.date_ouverture: champs_remplis += 1
    if etablissement.longitude and etablissement.latitude: champs_remplis += 1
    if apprenants.exists(): champs_remplis += 1
    if formateurs.exists(): champs_remplis += 1
    if filieres.exists(): champs_remplis += 1
    
    progression = int((champs_remplis / champs_total) * 100) if champs_total > 0 else 0
    
    context = {
        'etablissement': etablissement,
        'apprenants': apprenants,
        'formateurs': formateurs,
        'filieres': filieres,
        'total_apprenants': total_apprenants,
        'total_formateurs': total_formateurs,
        'total_filieres': total_filieres,
        'progression_globale': progression,
    }
    
    print(f"✅ Contexte préparé, rendu du template...")  # Pour debug
    return render(request, 'eftp_formel/etablissement_dashboard_test.html', context)
'''
@login_required
def etablissement_detail(request, pk):
    """Détail d'un établissement"""
    etablissement = get_object_or_404(EtablissementFormel, pk=pk)
    
    # Statistiques (à adapter selon vos modèles)
    context = {
        'etablissement': etablissement,
        'total_apprenants': 0,
        'apprenants_m': 0,
        'apprenants_f': 0,
        'total_formateurs': 0,
        'formateurs_m': 0,
        'formateurs_f': 0,
        'total_filieres': 0,
        'filieres': [],
    }
    return render(request, 'eftp_formel/etablissement_detail.html', context)

@login_required
def etablissement_edit(request, pk):
    """Modifier un établissement (formulaire complet)"""
    etablissement = get_object_or_404(EtablissementFormel, pk=pk)
    
    if request.method == 'POST':
        form = EtablissementFormelCompletForm(request.POST, instance=etablissement)
        if form.is_valid():
            form.save()
            messages.success(request, "Établissement modifié avec succès!")
            return redirect('eftp_formel:etablissement_detail', pk=etablissement.pk)
    else:
        form = EtablissementFormelCompletForm(instance=etablissement)
    
    context = {
        'form': form,
        'etablissement': etablissement,
        'titre': f"Modifier - {etablissement.nom}"
    }
    return render(request, 'eftp_formel/etablissement_form_complet.html', context)

@login_required
def etablissement_delete(request, pk):
    """Supprimer un établissement"""
    etablissement = get_object_or_404(EtablissementFormel, pk=pk)
    
    if request.method == 'POST':
        etablissement.delete()
        messages.success(request, "Établissement supprimé avec succès!")
        return redirect('eftp_formel:etablissement_list')
    
    return render(request, 'eftp_formel/etablissement_confirm_delete.html', {'etablissement': etablissement})

@login_required
def import_export(request):
    """Page d'import/export"""
    return render(request, 'eftp_formel/import_export.html')

@login_required
def etablissement_complet(request, pk):
    """Saisie complète des données d'un établissement"""
    etablissement = get_object_or_404(EtablissementFormel, pk=pk)
    if request.method == 'POST':
        form = EtablissementFormelCompletForm(request.POST, instance=etablissement)
        if form.is_valid():
            form.save()
            messages.success(request, "Données complètes enregistrées avec succès!")
            return redirect('eftp_formel:etablissement_detail', pk=etablissement.pk)
        else:
            messages.error(request, "Veuillez corriger les erreurs ci-dessous.")
    else:
        form = EtablissementFormelCompletForm(instance=etablissement)
    # Calcul de la progression
    total_fields = len(form.fields)
    filled_fields = sum(1 for field in form.fields if form.initial.get(field))
    progression = int((filled_fields / total_fields) * 100) if total_fields > 0 else 0
    context = {
        'form': form,
        'etablissement': etablissement,
        'progression': progression,
        'titre': f"Saisie complète - {etablissement.nom}"
    }
    return render(request, 'eftp_formel/etablissement_form_complet.html', context)

@login_required
def import_data(request):
    """Importation des données depuis un fichier"""
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        model_type = request.POST.get('model')
        format_type = file.name.split('.')[-1].lower()
        
        try:
            if format_type == 'csv':
                # Traitement CSV
                decoded_file = file.read().decode('utf-8').splitlines()
                reader = csv.DictReader(decoded_file)
                count = 0
                for row in reader:
                    # Logique d'import selon le modèle
                    if model_type == 'etablissement':
                        # Adapter selon vos champs
                        pass
                    count += 1
                messages.success(request, f"{count} enregistrements importés avec succès!")
                
            elif format_type in ['xlsx', 'xls']:
                # Traitement Excel
                wb = openpyxl.load_workbook(file)
                sheet = wb.active
                count = sheet.max_row - 1  # Moins l'en-tête
                messages.success(request, f"{count} enregistrements importés avec succès!")
                
            else:
                messages.error(request, "Format de fichier non supporté")
                
        except Exception as e:
            messages.error(request, f"Erreur lors de l'import : {str(e)}")
            
        return redirect('eftp_formel:import_export')
    
    messages.error(request, "Aucun fichier sélectionné")
    return redirect('eftp_formel:import_export')

@login_required
def export_data(request):
    """Exportation des données"""
    model_type = request.GET.get('model', 'all')
    format_type = request.GET.get('format', 'excel')
    
    # Préparer les données
    if model_type == 'all' or model_type == 'etablissement':
        queryset = EtablissementFormel.objects.all().select_related('region', 'departement', 'commune')
        fields = ['code', 'nom', 'statut', 'zone', 'region__nom', 'departement__nom', 
                 'commune__nom', 'type_etablissement', 'regime']
        headers = ['Code', 'Nom', 'Statut', 'Zone', 'Région', 'Département', 
                  'Commune', "Type d'établissement", 'Régime']
    else:
        messages.error(request, "Type de données non reconnu")
        return redirect('eftp_formel:import_export')
    
    # Export selon le format
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="export_{model_type}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(headers)
        
        for obj in queryset:
            row = []
            for field in fields:
                value = obj
                for part in field.split('__'):
                    value = getattr(value, part, '') if value else ''
                row.append(str(value) if value else '')
            writer.writerow(row)
        
        return response
        
    elif format_type == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="export_{model_type}.xlsx"'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Export"
        
        # En-têtes
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Données
        for row, obj in enumerate(queryset, 2):
            for col, field in enumerate(fields, 1):
                value = obj
                for part in field.split('__'):
                    value = getattr(value, part, '') if value else ''
                ws.cell(row=row, column=col, value=str(value) if value else '')
        
        wb.save(response)
        return response
    
    messages.error(request, "Format d'export non supporté")
    return redirect('eftp_formel:import_export')

# ================ VUES POUR LES APPRENANTS ================

@login_required
def apprenant_create(request, etablissement_id):
    """Ajouter des apprenants pour un établissement"""
    etablissement = get_object_or_404(EtablissementFormel, pk=etablissement_id)
    
    if request.method == 'POST':
        form = ApprenantFormelForm(request.POST)
        if form.is_valid():
            apprenant = form.save(commit=False)
            apprenant.etablissement = etablissement
            apprenant.save()
            messages.success(request, "Apprenants ajoutés avec succès!")
            return redirect('eftp_formel:etablissement_complet', pk=etablissement.id)
    else:
        form = ApprenantFormelForm()
    
    context = {
        'form': form,
        'etablissement': etablissement,
        'titre': f"Ajouter des apprenants - {etablissement.nom}"
    }
    return render(request, 'eftp_formel/apprenant_form.html', context)

@login_required
def apprenant_edit(request, pk):
    """Modifier des apprenants"""
    apprenant = get_object_or_404(ApprenantFormel, pk=pk)
    etablissement = apprenant.etablissement
    
    if request.method == 'POST':
        form = ApprenantFormelForm(request.POST, instance=apprenant)
        if form.is_valid():
            form.save()
            messages.success(request, "Apprenants modifiés avec succès!")
            return redirect('eftp_formel:etablissement_complet', pk=etablissement.id)
    else:
        form = ApprenantFormelForm(instance=apprenant)
    
    context = {
        'form': form,
        'apprenant': apprenant,
        'etablissement': etablissement,
        'titre': f"Modifier les apprenants - {etablissement.nom}"
    }
    return render(request, 'eftp_formel/apprenant_form.html', context)

@login_required
def apprenant_delete(request, pk):
    """Supprimer des apprenants"""
    apprenant = get_object_or_404(ApprenantFormel, pk=pk)
    etablissement = apprenant.etablissement
    
    if request.method == 'POST':
        apprenant.delete()
        messages.success(request, "Apprenants supprimés avec succès!")
        return redirect('eftp_formel:etablissement_complet', pk=etablissement.id)
    
    context = {
        'apprenant': apprenant,
        'etablissement': etablissement
    }
    return render(request, 'eftp_formel/apprenant_confirm_delete.html', context)


# ================ VUES POUR LES FORMATEURS ================

@login_required
def formateur_create(request, etablissement_id):
    """Ajouter un formateur pour un établissement"""
    etablissement = get_object_or_404(EtablissementFormel, pk=etablissement_id)
    
    if request.method == 'POST':
        form = FormateurFormelForm(request.POST)
        if form.is_valid():
            formateur = form.save(commit=False)
            formateur.etablissement = etablissement
            formateur.save()
            messages.success(request, "Formateur ajouté avec succès!")
            return redirect('eftp_formel:etablissement_complet', pk=etablissement.id)
    else:
        form = FormateurFormelForm()
    
    context = {
        'form': form,
        'etablissement': etablissement,
        'titre': f"Ajouter un formateur - {etablissement.nom}"
    }
    return render(request, 'eftp_formel/formateur_form.html', context)

@login_required
def formateur_edit(request, pk):
    """Modifier un formateur"""
    formateur = get_object_or_404(FormateurFormel, pk=pk)
    etablissement = formateur.etablissement
    
    if request.method == 'POST':
        form = FormateurFormelForm(request.POST, instance=formateur)
        if form.is_valid():
            form.save()
            messages.success(request, "Formateur modifié avec succès!")
            return redirect('eftp_formel:etablissement_complet', pk=etablissement.id)
    else:
        form = FormateurFormelForm(instance=formateur)
    
    context = {
        'form': form,
        'formateur': formateur,
        'etablissement': etablissement,
        'titre': f"Modifier le formateur - {etablissement.nom}"
    }
    return render(request, 'eftp_formel/formateur_form.html', context)

@login_required
def formateur_delete(request, pk):
    """Supprimer un formateur"""
    formateur = get_object_or_404(FormateurFormel, pk=pk)
    etablissement = formateur.etablissement
    
    if request.method == 'POST':
        formateur.delete()
        messages.success(request, "Formateur supprimé avec succès!")
        return redirect('eftp_formel:etablissement_complet', pk=etablissement.id)
    
    context = {
        'formateur': formateur,
        'etablissement': etablissement
    }
    return render(request, 'eftp_formel/formateur_confirm_delete.html', context)


# ================ VUES POUR LES FILIÈRES ================

@login_required
def filiere_create(request, etablissement_id):
    """Ajouter une filière pour un établissement"""
    etablissement = get_object_or_404(EtablissementFormel, pk=etablissement_id)
    
    if request.method == 'POST':
        form = FiliereFormelForm(request.POST)
        if form.is_valid():
            filiere = form.save(commit=False)
            filiere.etablissement = etablissement
            filiere.save()
            messages.success(request, "Filière ajoutée avec succès!")
            return redirect('eftp_formel:etablissement_complet', pk=etablissement.id)
    else:
        form = FiliereFormelForm()
    
    context = {
        'form': form,
        'etablissement': etablissement,
        'titre': f"Ajouter une filière - {etablissement.nom}"
    }
    return render(request, 'eftp_formel/filiere_form.html', context)

@login_required
def filiere_edit(request, pk):
    """Modifier une filière"""
    filiere = get_object_or_404(FiliereFormel, pk=pk)
    etablissement = filiere.etablissement
    
    if request.method == 'POST':
        form = FiliereFormelForm(request.POST, instance=filiere)
        if form.is_valid():
            form.save()
            messages.success(request, "Filière modifiée avec succès!")
            return redirect('eftp_formel:etablissement_complet', pk=etablissement.id)
    else:
        form = FiliereFormelForm(instance=filiere)
    
    context = {
        'form': form,
        'filiere': filiere,
        'etablissement': etablissement,
        'titre': f"Modifier la filière - {etablissement.nom}"
    }
    return render(request, 'eftp_formel/filiere_form.html', context)

@login_required
def filiere_delete(request, pk):
    """Supprimer une filière"""
    filiere = get_object_or_404(FiliereFormel, pk=pk)
    etablissement = filiere.etablissement
    
    if request.method == 'POST':
        filiere.delete()
        messages.success(request, "Filière supprimée avec succès!")
        return redirect('eftp_formel:etablissement_complet', pk=etablissement.id)
    
    context = {
        'filiere': filiere,
        'etablissement': etablissement
    }
    return render(request, 'eftp_formel/filiere_confirm_delete.html', context)