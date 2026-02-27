#!/usr/bin/env python
"""
Script pour générer tous les modèles de fichiers Excel pour l'import/export
"""
import os
import django
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Configuration Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from apps.renaloc.models import Region, Departement, Commune
from apps.eftp_formel.models import EtablissementFormel

def create_excel_template(filename, title, headers, sheet_name="Modèle", descriptions=None):
    """Crée un fichier Excel modèle avec les en-têtes et des exemples"""
    
    # Créer le workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    example_font = Font(color="666666", italic=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Créer les en-têtes
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Ajouter une ligne d'exemple si des descriptions sont fournies
    if descriptions:
        for col, desc in enumerate(descriptions, 1):
            cell = ws.cell(row=2, column=col, value=desc)
            cell.font = example_font
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = border
    
    # Ajouter une note explicative
    ws.cell(row=4, column=1, value="INSTRUCTIONS:").font = Font(bold=True)
    ws.cell(row=5, column=1, value="1. Ne pas modifier la première ligne (en-têtes)")
    ws.cell(row=6, column=1, value="2. Remplir à partir de la ligne 2")
    ws.cell(row=7, column=1, value="3. Les champs marqués * sont obligatoires")
    ws.cell(row=8, column=1, value="4. Respecter le format des dates (JJ/MM/AAAA)")
    
    # Ajuster la largeur des colonnes
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    
    # Sauvegarder
    os.makedirs('static/templates', exist_ok=True)
    filepath = f'static/templates/{filename}'
    wb.save(filepath)
    print(f"✅ Modèle créé : {filepath}")
    
    return filepath

def generate_all_templates():
    """Génère tous les modèles Excel"""
    
    print("🚀 Génération des modèles de fichiers...\n")
    
    # 1. Modèle Établissements
    etablissement_headers = [
        'Code*', 'Nom*', 'Sigle', 'Statut*', 'Zone*', 'Région*', 
        'Département*', 'Commune*', "Type d'établissement*", 'Régime*',
        'Date autorisation', "Date d'ouverture", 'Adresse', 'Longitude', 'Latitude'
    ]
    etablissement_example = [
        'LP001', 'Lycée Professionnel de Niamey', 'LP-NIA', 'PUBLIC', 'URBAINE',
        'Niamey', 'Niamey', 'Niamey', 'LP', 'EXTERNAT',
        '01/01/2020', '15/09/2020', 'Quartier Plateau', '2.123456', '13.512345'
    ]
    create_excel_template(
        'etablissement_template.xlsx',
        'Modèle Établissements',
        etablissement_headers,
        'Établissements',
        etablissement_example
    )
    
    # 2. Modèle Apprenants
    apprenant_headers = [
        'Code Établissement*', 'Cycle*', 'Année étude*', 
        'Masculin', 'Féminin', 'Redoublants M', 'Redoublants F'
    ]
    apprenant_example = [
        'LP001', 'BASE_1', '1ERE', '25', '15', '2', '1'
    ]
    create_excel_template(
        'apprenant_template.xlsx',
        'Modèle Apprenants',
        apprenant_headers,
        'Apprenants',
        apprenant_example
    )
    
    # 3. Modèle Filières
    filiere_headers = [
        'Code Établissement*', 'Secteur*', 'Nom filière*', 'Diplôme préparé*',
        'Cycle*', 'Durée formation (mois)*', 'Effectif M', 'Effectif F',
        'Stage obligatoire (OUI/NON)', 'Heures pratique/semaine'
    ]
    filiere_example = [
        'LP001', 'TERTIAIRE', 'Comptabilité et Gestion', 'BTS Comptabilité',
        'MOYEN_2', '24', '18', '22', 'OUI', '15'
    ]
    create_excel_template(
        'filiere_template.xlsx',
        'Modèle Filières',
        filiere_headers,
        'Filières',
        filiere_example
    )
    
    # 4. Modèle Formateurs
    formateur_headers = [
        'Code Établissement*', 'Nom et prénom*', 'Sexe (M/F)*', 'Date naissance*',
        'Année recrutement*', 'Statut*', 'Nationalité*', 'Diplôme académique',
        'Diplôme professionnel', 'Disciplines enseignées*', 'Volume horaire/semaine*'
    ]
    formateur_example = [
        'LP001', 'SANI Issoufou', 'M', '15/05/1985',
        '2010', 'FONCTIONNAIRE', 'NIGERIENNE', 'Master en Mathématiques',
        'CAPET', 'Mathématiques, Physique', '18'
    ]
    create_excel_template(
        'formateur_template.xlsx',
        'Modèle Formateurs',
        formateur_headers,
        'Formateurs',
        formateur_example
    )
    
    # 5. Modèle pour les valeurs possibles (aide)
    valeurs_headers = [
        'Champ', 'Valeurs possibles', 'Description'
    ]
    valeurs_data = [
        ['Statut', 'PUBLIC, PRIVE', 'Statut de l\'établissement'],
        ['Zone', 'URBAINE, RURALE', 'Zone géographique'],
        ['Cycle', 'BASE_1, BASE_2, MOYEN_1, MOYEN_2', 'Cycle de formation'],
        ['Secteur', 'PRIMAIRE, SECONDAIRE, TERTIAIRE', 'Secteur d\'activité'],
        ['Type établissement', 'LP, LT, LTE, LA, CFPT, CFPP, CMCAN, EI, CET, CFM, CPJ, EFAC, CFPT_AMA, ENI, CENTRE_MUSEE, CFJA, CFMAA, AUTRE', 'Type d\'établissement'],
        ['Régime', 'INTERNAT, EXTERNAT', 'Régime de l\'établissement'],
        ['Statut formateur', 'FONCTIONNAIRE, CONTRACTUEL, ASCN, PERMANENT, VOLONTAIRE', 'Statut du formateur'],
        ['Nationalité', 'NIGERIENNE, ETRANGERE', 'Nationalité'],
        ['Oui/Non', 'OUI, NON', 'Pour les champs booléens'],
    ]
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Valeurs possibles"
    
    # Style pour les en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # En-têtes
    for col, header in enumerate(valeurs_headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Données
    for row, data in enumerate(valeurs_data, 2):
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Ajuster les largeurs
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 30
    
    wb.save('static/templates/valeurs_reference.xlsx')
    print("✅ Modèle créé : static/templates/valeurs_reference.xlsx")
    
    print("\n🎉 Tous les modèles ont été générés avec succès!")
    print("📁 Emplacement : static/templates/")

if __name__ == "__main__":
    generate_all_templates()